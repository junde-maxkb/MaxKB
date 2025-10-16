# coding=utf-8
"""
    @project: maxkb
    @Author：虎
    @file： document_serializers.py
    @date：2023/9/22 13:43
    @desc:
"""
import io
import logging
import os
import re
import traceback
import uuid
from functools import reduce
from tempfile import TemporaryDirectory
from typing import List, Dict

import openpyxl
from celery_once import AlreadyQueued
from django.core import validators
from django.db import transaction, models
from django.db.models import QuerySet, Count
from django.db.models.functions import Substr, Reverse
from django.http import HttpResponse
from drf_yasg import openapi
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from rest_framework import serializers
from xlwt import Utils

from common.db.search import native_search, native_page_search, get_dynamics_model
from common.event import ListenerManagement
from common.event.common import work_thread_pool
from common.exception.app_exception import AppApiException
from common.handle.impl.csv_split_handle import CsvSplitHandle
from common.handle.impl.doc_split_handle import DocSplitHandle
from common.handle.impl.html_split_handle import HTMLSplitHandle
from common.handle.impl.pdf_split_handle import PdfSplitHandle
from common.handle.impl.qa.csv_parse_qa_handle import CsvParseQAHandle
from common.handle.impl.qa.xls_parse_qa_handle import XlsParseQAHandle
from common.handle.impl.qa.xlsx_parse_qa_handle import XlsxParseQAHandle
from common.handle.impl.qa.zip_parse_qa_handle import ZipParseQAHandle
from common.handle.impl.table.csv_parse_table_handle import CsvSplitHandle as CsvSplitTableHandle
from common.handle.impl.table.xls_parse_table_handle import XlsSplitHandle as XlsSplitTableHandle
from common.handle.impl.table.xlsx_parse_table_handle import XlsxSplitHandle as XlsxSplitTableHandle
from common.handle.impl.text_split_handle import TextSplitHandle
from common.handle.impl.xls_split_handle import XlsSplitHandle
from common.handle.impl.xlsx_split_handle import XlsxSplitHandle
from common.handle.impl.zip_split_handle import ZipSplitHandle
from common.mixins.api_mixin import ApiMixin
from common.util.common import post, flat_map, bulk_create_in_batches, parse_image
from common.util.field_message import ErrMessage
from common.util.file_util import get_file_content
from common.util.fork import Fork
from common.util.split_model import get_split_model
from dataset.models.data_set import DataSet, Document, Paragraph, Problem, Type, ProblemParagraphMapping, Image, \
    TaskType, State
from dataset.serializers.common_serializers import BatchSerializer, MetaSerializer, ProblemParagraphManage, \
    get_embedding_model_id_by_dataset_id, write_image, zip_dir
from dataset.serializers.paragraph_serializers import ParagraphSerializers, ParagraphInstanceSerializer
from dataset.task import sync_web_document, generate_related_by_document_id
from embedding.task.embedding import embedding_by_document, delete_embedding_by_document_list, \
    delete_embedding_by_document, update_embedding_dataset_id, delete_embedding_by_paragraph_ids, \
    embedding_by_document_list
from setting.models import Model
from setting.models.data_source import DataSourceConfig
from setting.models_provider.tools import DBConnector
from smartdoc.conf import PROJECT_DIR
from django.utils.translation import gettext_lazy as _, gettext, to_locale
from django.utils.translation import get_language

parse_qa_handle_list = [XlsParseQAHandle(), CsvParseQAHandle(), XlsxParseQAHandle(), ZipParseQAHandle()]
parse_table_handle_list = [CsvSplitTableHandle(), XlsSplitTableHandle(), XlsxSplitTableHandle()]


class FileBufferHandle:
    buffer = None

    def get_buffer(self, file):
        if self.buffer is None:
            self.buffer = file.read()
        return self.buffer


class BatchCancelInstanceSerializer(serializers.Serializer):
    id_list = serializers.ListField(required=True, child=serializers.UUIDField(required=True),
                                    error_messages=ErrMessage.char(_('id list')))
    type = serializers.IntegerField(required=True, error_messages=ErrMessage.integer(
        _('task type')))

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)
        _type = self.data.get('type')
        try:
            TaskType(_type)
        except Exception as e:
            raise AppApiException(500, _('task type not support'))


class CancelInstanceSerializer(serializers.Serializer):
    type = serializers.IntegerField(required=True, error_messages=ErrMessage.integer(
        _('task type')))

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)
        _type = self.data.get('type')
        try:
            TaskType(_type)
        except Exception as e:
            raise AppApiException(500, _('task type not support'))


class DocumentEditInstanceSerializer(ApiMixin, serializers.Serializer):
    meta = serializers.DictField(required=False)
    name = serializers.CharField(required=False, max_length=128, min_length=1,
                                 error_messages=ErrMessage.char(
                                     _('document name')))
    hit_handling_method = serializers.CharField(required=False, validators=[
        validators.RegexValidator(regex=re.compile("^optimization|directly_return$"),
                                  message=_('The type only supports optimization|directly_return'),
                                  code=500)
    ], error_messages=ErrMessage.char(_('hit handling method')))

    directly_return_similarity = serializers.FloatField(required=False,
                                                        max_value=2,
                                                        min_value=0,
                                                        error_messages=ErrMessage.float(
                                                            _('directly return similarity')))

    is_active = serializers.BooleanField(required=False, error_messages=ErrMessage.boolean(
        _('document is active')))

    @staticmethod
    def get_meta_valid_map():
        dataset_meta_valid_map = {
            Type.base: MetaSerializer.BaseMeta,
            Type.web: MetaSerializer.WebMeta
        }
        return dataset_meta_valid_map

    def is_valid(self, *, document: Document = None):
        super().is_valid(raise_exception=True)
        if 'meta' in self.data and self.data.get('meta') is not None:
            dataset_meta_valid_map = self.get_meta_valid_map()
            valid_class = dataset_meta_valid_map.get(document.type)
            valid_class(data=self.data.get('meta')).is_valid(raise_exception=True)


class DocumentWebInstanceSerializer(ApiMixin, serializers.Serializer):
    source_url_list = serializers.ListField(required=True,
                                            child=serializers.CharField(required=True, error_messages=ErrMessage.char(
                                                _('document url list'))),
                                            error_messages=ErrMessage.char(
                                                _('document url list')))
    selector = serializers.CharField(required=False, allow_null=True, allow_blank=True,
                                     error_messages=ErrMessage.char(
                                         _('selector')))

    @staticmethod
    def get_request_params_api():
        return [openapi.Parameter(name='file',
                                  in_=openapi.IN_FORM,
                                  type=openapi.TYPE_ARRAY,
                                  items=openapi.Items(type=openapi.TYPE_FILE),
                                  required=True,
                                  description=_('file')),
                openapi.Parameter(name='dataset_id',
                                  in_=openapi.IN_PATH,
                                  type=openapi.TYPE_STRING,
                                  required=True,
                                  description=_('dataset id')),
                ]

    @staticmethod
    def get_request_body_api():
        return openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['source_url_list'],
            properties={
                'source_url_list': openapi.Schema(type=openapi.TYPE_ARRAY, title=_('source url list'),
                                                  description=_('source url list'),
                                                  items=openapi.Schema(type=openapi.TYPE_STRING)),
                'selector': openapi.Schema(type=openapi.TYPE_STRING, title=_('selector'), description=_('selector'))
            }
        )


class DocumentInstanceSerializer(ApiMixin, serializers.Serializer):
    name = serializers.CharField(required=True,
                                 error_messages=ErrMessage.char(_('document name')),
                                 max_length=128,
                                 min_length=1)

    paragraphs = ParagraphInstanceSerializer(required=False, many=True, allow_null=True)

    @staticmethod
    def get_request_body_api():
        return openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['name', 'paragraphs'],
            properties={
                'name': openapi.Schema(type=openapi.TYPE_STRING, title=_('document name'),
                                       description=_('document name')),
                'paragraphs': openapi.Schema(type=openapi.TYPE_ARRAY, title=_('paragraphs'),
                                             description=_('paragraphs'),
                                             items=ParagraphSerializers.Create.get_request_body_api())
            }
        )


class DocumentInstanceQASerializer(ApiMixin, serializers.Serializer):
    file_list = serializers.ListSerializer(required=True,
                                           error_messages=ErrMessage.list(_('file list')),
                                           child=serializers.FileField(required=True,
                                                                       error_messages=ErrMessage.file(_('file'))))


class DocumentInstanceTableSerializer(ApiMixin, serializers.Serializer):
    file_list = serializers.ListSerializer(required=True,
                                           error_messages=ErrMessage.list(_('file list')),
                                           child=serializers.FileField(required=True,
                                                                       error_messages=ErrMessage.file(_('file'))))


class DocumentSerializers(ApiMixin, serializers.Serializer):
    class Export(ApiMixin, serializers.Serializer):
        type = serializers.CharField(required=True, validators=[
            validators.RegexValidator(regex=re.compile("^csv|excel$"),
                                      message=_('The template type only supports excel|csv'),
                                      code=500)
        ], error_messages=ErrMessage.char(_('type')))

        @staticmethod
        def get_request_params_api():
            return [openapi.Parameter(name='type',
                                      in_=openapi.IN_QUERY,
                                      type=openapi.TYPE_STRING,
                                      required=True,
                                      description=_('Export template type csv|excel')),

                    ]

        def export(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            language = get_language()
            if self.data.get('type') == 'csv':
                file = open(
                    os.path.join(PROJECT_DIR, "apps", "dataset", 'template', f'csv_template_{to_locale(language)}.csv'),
                    "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'text/csv',
                                                                  'Content-Disposition': 'attachment; filename="csv_template.csv"'})
            elif self.data.get('type') == 'excel':
                file = open(os.path.join(PROJECT_DIR, "apps", "dataset", 'template',
                                         f'excel_template_{to_locale(language)}.xlsx'), "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'application/vnd.ms-excel',
                                                                  'Content-Disposition': 'attachment; filename="excel_template.xlsx"'})

        def table_export(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            language = get_language()
            if self.data.get('type') == 'csv':
                file = open(
                    os.path.join(PROJECT_DIR, "apps", "dataset", 'template',
                                 f'table_template_{to_locale(language)}.csv'),
                    "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'text/cxv',
                                                                  'Content-Disposition': 'attachment; filename="csv_template.csv"'})
            elif self.data.get('type') == 'excel':
                file = open(os.path.join(PROJECT_DIR, "apps", "dataset", 'template',
                                         f'table_template_{to_locale(language)}.xlsx'),
                            "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'application/vnd.ms-excel',
                                                                  'Content-Disposition': 'attachment; filename="excel_template.xlsx"'})

    class Migrate(ApiMixin, serializers.Serializer):
        dataset_id = serializers.UUIDField(required=True,
                                           error_messages=ErrMessage.char(
                                               _('dataset id')))
        target_dataset_id = serializers.UUIDField(required=True,
                                                  error_messages=ErrMessage.char(
                                                      _('target dataset id')))
        document_id_list = serializers.ListField(required=True, error_messages=ErrMessage.char(_('document list')),
                                                 child=serializers.UUIDField(required=True,
                                                                             error_messages=ErrMessage.uuid(
                                                                                 _('document id'))))

        @transaction.atomic
        def migrate(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            dataset_id = self.data.get('dataset_id')
            target_dataset_id = self.data.get('target_dataset_id')
            dataset = QuerySet(DataSet).filter(id=dataset_id).first()
            target_dataset = QuerySet(DataSet).filter(id=target_dataset_id).first()
            document_id_list = self.data.get('document_id_list')
            document_list = QuerySet(Document).filter(dataset_id=dataset_id, id__in=document_id_list)
            paragraph_list = QuerySet(Paragraph).filter(dataset_id=dataset_id, document_id__in=document_id_list)

            problem_paragraph_mapping_list = QuerySet(ProblemParagraphMapping).filter(paragraph__in=paragraph_list)
            problem_list = QuerySet(Problem).filter(
                id__in=[problem_paragraph_mapping.problem_id for problem_paragraph_mapping in
                        problem_paragraph_mapping_list])
            target_problem_list = list(
                QuerySet(Problem).filter(content__in=[problem.content for problem in problem_list],
                                         dataset_id=target_dataset_id))
            target_handle_problem_list = [
                self.get_target_dataset_problem(target_dataset_id, problem_paragraph_mapping,
                                                problem_list, target_problem_list) for
                problem_paragraph_mapping
                in
                problem_paragraph_mapping_list]

            create_problem_list = [problem for problem, is_create in target_handle_problem_list if
                                   is_create is not None and is_create]
            # 插入问题
            QuerySet(Problem).bulk_create(create_problem_list)
            # 修改mapping
            QuerySet(ProblemParagraphMapping).bulk_update(problem_paragraph_mapping_list, ['problem_id', 'dataset_id'])
            # 修改文档
            if dataset.type == Type.base.value and target_dataset.type == Type.web.value:
                document_list.update(dataset_id=target_dataset_id, type=Type.web,
                                     meta={'source_url': '', 'selector': ''})
            elif target_dataset.type == Type.base.value and dataset.type == Type.web.value:
                document_list.update(dataset_id=target_dataset_id, type=Type.base,
                                     meta={})
            else:
                document_list.update(dataset_id=target_dataset_id)
            model_id = None
            if dataset.embedding_mode_id != target_dataset.embedding_mode_id:
                model_id = get_embedding_model_id_by_dataset_id(target_dataset_id)

            pid_list = [paragraph.id for paragraph in paragraph_list]
            # 修改段落信息
            paragraph_list.update(dataset_id=target_dataset_id)
            # 修改向量信息
            if model_id:
                delete_embedding_by_paragraph_ids(pid_list)
                ListenerManagement.update_status(QuerySet(Document).filter(id__in=document_id_list),
                                                 TaskType.EMBEDDING,
                                                 State.PENDING)
                ListenerManagement.update_status(QuerySet(Paragraph).filter(document_id__in=document_id_list),
                                                 TaskType.EMBEDDING,
                                                 State.PENDING)
                ListenerManagement.get_aggregation_document_status_by_query_set(
                    QuerySet(Document).filter(id__in=document_id_list))()
                embedding_by_document_list.delay(document_id_list, model_id)
            else:
                update_embedding_dataset_id(pid_list, target_dataset_id)

        @staticmethod
        def get_target_dataset_problem(target_dataset_id: str,
                                       problem_paragraph_mapping,
                                       source_problem_list,
                                       target_problem_list):
            source_problem_list = [source_problem for source_problem in source_problem_list if
                                   source_problem.id == problem_paragraph_mapping.problem_id]
            problem_paragraph_mapping.dataset_id = target_dataset_id
            if len(source_problem_list) > 0:
                problem_content = source_problem_list[-1].content
                problem_list = [problem for problem in target_problem_list if problem.content == problem_content]
                if len(problem_list) > 0:
                    problem = problem_list[-1]
                    problem_paragraph_mapping.problem_id = problem.id
                    return problem, False
                else:
                    problem = Problem(id=uuid.uuid1(), dataset_id=target_dataset_id, content=problem_content)
                    target_problem_list.append(problem)
                    problem_paragraph_mapping.problem_id = problem.id
                    return problem, True
            return None

        @staticmethod
        def get_request_params_api():
            return [openapi.Parameter(name='dataset_id',
                                      in_=openapi.IN_PATH,
                                      type=openapi.TYPE_STRING,
                                      required=True,
                                      description=_('document id')),
                    openapi.Parameter(name='target_dataset_id',
                                      in_=openapi.IN_PATH,
                                      type=openapi.TYPE_STRING,
                                      required=True,
                                      description=_('target document id'))
                    ]

        @staticmethod
        def get_request_body_api():
            return openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(type=openapi.TYPE_STRING),
                title=_('document id list'),
                description=_('document id list')
            )

    class Query(ApiMixin, serializers.Serializer):
        # 知识库id
        dataset_id = serializers.UUIDField(required=True,
                                           error_messages=ErrMessage.char(
                                               _('dataset id')))

        name = serializers.CharField(required=False, max_length=128,
                                     min_length=1,
                                     error_messages=ErrMessage.char(
                                         _('document name')))
        hit_handling_method = serializers.CharField(required=False,
                                                    error_messages=ErrMessage.char(_('hit handling method')))
        is_active = serializers.BooleanField(required=False, error_messages=ErrMessage.boolean(_('document is active')))
        task_type = serializers.IntegerField(required=False, error_messages=ErrMessage.integer(_('task type')))
        status = serializers.CharField(required=False, error_messages=ErrMessage.char(_('status')))
        order_by = serializers.CharField(required=False, error_messages=ErrMessage.char(_('order by')))

        def get_query_set(self):
            query_set = QuerySet(model=Document)
            query_set = query_set.filter(**{'dataset_id': self.data.get("dataset_id")})
            if 'name' in self.data and self.data.get('name') is not None:
                query_set = query_set.filter(**{'name__icontains': self.data.get('name')})
            if 'hit_handling_method' in self.data and self.data.get('hit_handling_method') is not None:
                query_set = query_set.filter(**{'hit_handling_method': self.data.get('hit_handling_method')})
            if 'is_active' in self.data and self.data.get('is_active') is not None:
                query_set = query_set.filter(**{'is_active': self.data.get('is_active')})
            if 'status' in self.data and self.data.get(
                    'status') is not None:
                task_type = self.data.get('task_type')
                status = self.data.get(
                    'status')
                if task_type is not None:
                    query_set = query_set.annotate(
                        reversed_status=Reverse('status'),
                        task_type_status=Substr('reversed_status', TaskType(task_type).value,
                                                1),
                    ).filter(task_type_status=State(status).value).values('id')
                else:
                    if status != State.SUCCESS.value:
                        query_set = query_set.filter(status__icontains=status)
                    else:
                        query_set = query_set.filter(status__iregex='^[2n]*$')
            order_by = self.data.get('order_by', '')
            order_by_query_set = QuerySet(model=get_dynamics_model(
                {'char_length': models.CharField(), 'paragraph_count': models.IntegerField(),
                 "update_time": models.IntegerField(), 'create_time': models.DateTimeField()}))
            if order_by:
                order_by_query_set = order_by_query_set.order_by(order_by)
            else:
                order_by_query_set = order_by_query_set.order_by('-create_time', 'id')
            return {
                'document_custom_sql': query_set,
                'order_by_query': order_by_query_set
            }

        def list(self, with_valid=False):
            if with_valid:
                self.is_valid(raise_exception=True)
            query_set = self.get_query_set()
            return native_search(query_set, select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "dataset", 'sql', 'list_document.sql')))

        def page(self, current_page, page_size):
            query_set = self.get_query_set()
            return native_page_search(current_page, page_size, query_set, select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "dataset", 'sql', 'list_document.sql')))

        @staticmethod
        def get_request_params_api():
            return [openapi.Parameter(name='name',
                                      in_=openapi.IN_QUERY,
                                      type=openapi.TYPE_STRING,
                                      required=False,
                                      description=_('document name')),
                    openapi.Parameter(name='hit_handling_method', in_=openapi.IN_QUERY,
                                      type=openapi.TYPE_STRING,
                                      required=False,
                                      description=_('hit handling method')), ]

        @staticmethod
        def get_response_body_api():
            return openapi.Schema(type=openapi.TYPE_ARRAY,
                                  title=_('document list'), description=_('document list'),
                                  items=DocumentSerializers.Operate.get_response_body_api())

    class Sync(ApiMixin, serializers.Serializer):
        document_id = serializers.UUIDField(required=True, error_messages=ErrMessage.char(
            _('document id')))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            document_id = self.data.get('document_id')
            first = QuerySet(Document).filter(id=document_id).first()
            if first is None:
                raise AppApiException(500, _('document id not exist'))
            if first.type != Type.web:
                raise AppApiException(500, _('Synchronization is only supported for web site types'))

        def sync(self, with_valid=True, with_embedding=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_id = self.data.get('document_id')
            document = QuerySet(Document).filter(id=document_id).first()
            state = State.SUCCESS
            if document.type != Type.web:
                return True
            try:
                ListenerManagement.update_status(QuerySet(Document).filter(id=document_id),
                                                 TaskType.SYNC,
                                                 State.PENDING)
                ListenerManagement.get_aggregation_document_status(document_id)()
                source_url = document.meta.get('source_url')
                selector_list = document.meta.get('selector').split(
                    " ") if 'selector' in document.meta and document.meta.get('selector') is not None else []
                result = Fork(source_url, selector_list).fork()
                if result.status == 200:
                    # 删除段落
                    QuerySet(model=Paragraph).filter(document_id=document_id).delete()
                    # 删除问题
                    QuerySet(model=ProblemParagraphMapping).filter(document_id=document_id).delete()
                    delete_problems_and_mappings([document_id])
                    # 删除向量库
                    delete_embedding_by_document(document_id)
                    paragraphs = get_split_model('web.md').parse(result.content)
                    char_length = reduce(lambda x, y: x + y,
                                         [len(p.get('content')) for p in paragraphs],
                                         0)
                    QuerySet(Document).filter(id=document_id).update(char_length=char_length)
                    document_paragraph_model = DocumentSerializers.Create.get_paragraph_model(document, paragraphs)

                    paragraph_model_list = document_paragraph_model.get('paragraph_model_list')
                    problem_paragraph_object_list = document_paragraph_model.get('problem_paragraph_object_list')
                    problem_model_list, problem_paragraph_mapping_list = ProblemParagraphManage(
                        problem_paragraph_object_list, document.dataset_id).to_problem_model_list()
                    # 批量插入段落
                    QuerySet(Paragraph).bulk_create(paragraph_model_list) if len(paragraph_model_list) > 0 else None
                    # 批量插入问题
                    QuerySet(Problem).bulk_create(problem_model_list) if len(problem_model_list) > 0 else None
                    # 插入关联问题
                    QuerySet(ProblemParagraphMapping).bulk_create(problem_paragraph_mapping_list) if len(
                        problem_paragraph_mapping_list) > 0 else None
                    # 向量化
                    if with_embedding:
                        embedding_model_id = get_embedding_model_id_by_dataset_id(document.dataset_id)
                        ListenerManagement.update_status(QuerySet(Document).filter(id=document_id),
                                                         TaskType.EMBEDDING,
                                                         State.PENDING)
                        ListenerManagement.update_status(QuerySet(Paragraph).filter(document_id=document_id),
                                                         TaskType.EMBEDDING,
                                                         State.PENDING)
                        ListenerManagement.get_aggregation_document_status(document_id)()
                        embedding_by_document.delay(document_id, embedding_model_id)

                else:
                    state = State.FAILURE
            except Exception as e:
                logging.getLogger("max_kb_error").error(f'{str(e)}:{traceback.format_exc()}')
                state = State.FAILURE
            ListenerManagement.update_status(QuerySet(Document).filter(id=document_id),
                                             TaskType.SYNC,
                                             state)
            ListenerManagement.update_status(QuerySet(Paragraph).filter(document_id=document_id),
                                             TaskType.SYNC,
                                             state)
            ListenerManagement.get_aggregation_document_status(document_id)()
            return True

    class Operate(ApiMixin, serializers.Serializer):
        document_id = serializers.UUIDField(required=True, error_messages=ErrMessage.char(
            _('document id')))
        dataset_id = serializers.UUIDField(required=True, error_messages=ErrMessage.char(_('dataset id')))

        @staticmethod
        def get_request_params_api():
            return [openapi.Parameter(name='dataset_id',
                                      in_=openapi.IN_PATH,
                                      type=openapi.TYPE_STRING,
                                      required=True,
                                      description=_('document id')),
                    openapi.Parameter(name='document_id',
                                      in_=openapi.IN_PATH,
                                      type=openapi.TYPE_STRING,
                                      required=True,
                                      description=_('document id'))
                    ]

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            document_id = self.data.get('document_id')
            if not QuerySet(Document).filter(id=document_id).exists():
                raise AppApiException(500, _('document id not exist'))

        def export(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document = QuerySet(Document).filter(id=self.data.get("document_id")).first()
            paragraph_list = native_search(QuerySet(Paragraph).filter(document_id=self.data.get("document_id")),
                                           get_file_content(
                                               os.path.join(PROJECT_DIR, "apps", "dataset", 'sql',
                                                            'list_paragraph_document_name.sql')))
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(document_id=self.data.get("document_id")), get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "dataset", 'sql', 'list_problem_mapping.sql')),
                with_table_name=True)
            data_dict, document_dict = self.merge_problem(paragraph_list, problem_mapping_list, [document])
            workbook = self.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="data.xlsx"'
            workbook.save(response)
            return response

        def export_zip(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document = QuerySet(Document).filter(id=self.data.get("document_id")).first()
            paragraph_list = native_search(QuerySet(Paragraph).filter(document_id=self.data.get("document_id")),
                                           get_file_content(
                                               os.path.join(PROJECT_DIR, "apps", "dataset", 'sql',
                                                            'list_paragraph_document_name.sql')))
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(document_id=self.data.get("document_id")), get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "dataset", 'sql', 'list_problem_mapping.sql')),
                with_table_name=True)
            data_dict, document_dict = self.merge_problem(paragraph_list, problem_mapping_list, [document])
            res = [parse_image(paragraph.get('content')) for paragraph in paragraph_list]

            workbook = DocumentSerializers.Operate.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="archive.zip"'
            zip_buffer = io.BytesIO()
            with TemporaryDirectory() as tempdir:
                dataset_file = os.path.join(tempdir, 'dataset.xlsx')
                workbook.save(dataset_file)
                for r in res:
                    write_image(tempdir, r)
                zip_dir(tempdir, zip_buffer)
            response.write(zip_buffer.getvalue())
            return response

        @staticmethod
        def get_workbook(data_dict, document_dict):
            # 创建工作簿对象
            workbook = openpyxl.Workbook()
            workbook.remove_sheet(workbook.active)
            if len(data_dict.keys()) == 0:
                data_dict['sheet'] = []
            for sheet_id in data_dict:
                # 添加工作表
                worksheet = workbook.create_sheet(document_dict.get(sheet_id))
                data = [
                    [gettext('Section title (optional)'),
                     gettext('Section content (required, question answer, no more than 4096 characters)'),
                     gettext('Question (optional, one per line in the cell)')],
                    *data_dict.get(sheet_id, [])
                ]
                # 写入数据到工作表
                for row_idx, row in enumerate(data):
                    for col_idx, col in enumerate(row):
                        cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        if isinstance(col, str):
                            col = re.sub(ILLEGAL_CHARACTERS_RE, '', col)
                        cell.value = col
                    # 创建HttpResponse对象返回Excel文件
            return workbook

        @staticmethod
        def merge_problem(paragraph_list: List[Dict], problem_mapping_list: List[Dict], document_list):
            result = {}
            document_dict = {}

            for paragraph in paragraph_list:
                problem_list = [problem_mapping.get('content') for problem_mapping in problem_mapping_list if
                                problem_mapping.get('paragraph_id') == paragraph.get('id')]
                document_sheet = result.get(paragraph.get('document_id'))
                document_name = DocumentSerializers.Operate.reset_document_name(paragraph.get('document_name'))
                d = document_dict.get(document_name)
                if d is None:
                    document_dict[document_name] = {paragraph.get('document_id')}
                else:
                    d.add(paragraph.get('document_id'))

                if document_sheet is None:
                    result[paragraph.get('document_id')] = [[paragraph.get('title'), paragraph.get('content'),
                                                             '\n'.join(problem_list)]]
                else:
                    document_sheet.append([paragraph.get('title'), paragraph.get('content'), '\n'.join(problem_list)])
            for document in document_list:
                if document.id not in result:
                    document_name = DocumentSerializers.Operate.reset_document_name(document.name)
                    result[document.id] = [[]]
                    d = document_dict.get(document_name)
                    if d is None:
                        document_dict[document_name] = {document.id}
                    else:
                        d.add(document.id)
            result_document_dict = {}
            for d_name in document_dict:
                for index, d_id in enumerate(document_dict.get(d_name)):
                    result_document_dict[d_id] = d_name if index == 0 else d_name + str(index)
            return result, result_document_dict

        @staticmethod
        def reset_document_name(document_name):
            if document_name is not None:
                document_name = document_name.strip()[0:29]
            if document_name is None or not Utils.valid_sheet_name(document_name):
                return "Sheet"
            return document_name.strip()

        def one(self, with_valid=False):
            if with_valid:
                self.is_valid(raise_exception=True)
            query_set = QuerySet(model=Document)
            query_set = query_set.filter(**{'id': self.data.get("document_id")})
            return native_search({
                'document_custom_sql': query_set,
                'order_by_query': QuerySet(Document).order_by('-create_time', 'id')
            }, select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "dataset", 'sql', 'list_document.sql')), with_search_one=True)

        def edit(self, instance: Dict, with_valid=False):
            if with_valid:
                self.is_valid(raise_exception=True)
            _document = QuerySet(Document).get(id=self.data.get("document_id"))
            if with_valid:
                DocumentEditInstanceSerializer(data=instance).is_valid(document=_document)
            update_keys = ['name', 'is_active', 'hit_handling_method', 'directly_return_similarity', 'meta']
            for update_key in update_keys:
                if update_key in instance and instance.get(update_key) is not None:
                    _document.__setattr__(update_key, instance.get(update_key))
            _document.save()
            return self.one()

        def refresh(self, state_list=None, with_valid=True):
            import logging
            logger = logging.getLogger('dataset.serializers.document_serializers')

            if state_list is None:
                state_list = [State.PENDING.value, State.STARTED.value, State.SUCCESS.value, State.FAILURE.value,
                              State.REVOKE.value,
                              State.REVOKED.value, State.IGNORED.value]
            if with_valid:
                self.is_valid(raise_exception=True)

            import datetime
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            logger.info(f"=== 开始文档向量化刷新调试信息 [{timestamp}] ===")
            logger.info(f"🎯 目标状态列表: {state_list}")

            dataset = QuerySet(DataSet).filter(id=self.data.get('dataset_id')).first()
            if dataset is None:
                logger.error("❌ 数据集不存在")
                raise AppApiException(500, _('Dataset does not exist'))

            logger.info(f"🗂️ 数据集ID: {dataset.id}")
            logger.info(f"📁 数据集名称: {dataset.name}")

            # 修复字段名错误：使用正确的字段名 embedding_mode_id (ForeignKey的id)
            embedding_model_id = dataset.embedding_mode_id
            dataset_user_id = dataset.user_id
            logger.info(f"🤖 向量模型ID: {embedding_model_id}")
            logger.info(f"👤 数据集用户ID: {dataset_user_id}")

            embedding_model = QuerySet(Model).filter(id=embedding_model_id).first()
            if embedding_model is None:
                logger.error("❌ 向量模型不存在")
                raise AppApiException(500, _('Model does not exist'))

            logger.info(f"🏷️ 向量模型名称: {embedding_model.name}")
            logger.info(f"🏢 向量模型提供商: {embedding_model.provider}")
            logger.info(f"🔐 向量模型权限类型: {embedding_model.permission_type}")

            if embedding_model.permission_type == 'PRIVATE' and dataset_user_id != embedding_model.user_id:
                logger.error(f"❌ 没有权限使用此模型: {embedding_model.name}")
                raise AppApiException(500, _('No permission to use this model') + f"{embedding_model.name}")

            document_id = self.data.get("document_id")
            logger.info(f"📄 文档ID: {document_id}")

            # 检查文档是否存在
            document = QuerySet(Document).filter(id=document_id).first()
            if document is None:
                logger.error("❌ 文档不存在")
                raise AppApiException(500, _('Document does not exist'))

            logger.info(f"📁 文档名称: {document.name}")
            logger.info(f"📋 文档类型: {document.type}")

            logger.info("🔄 更新文档向量化状态为PENDING...")
            ListenerManagement.update_status(QuerySet(Document).filter(id=document_id), TaskType.EMBEDDING,
                                             State.PENDING)

            # 修复EmptyResultSet错误：只有当存在符合条件的段落时才更新状态
            paragraph_queryset = QuerySet(Paragraph).annotate(
                reversed_status=Reverse('status'),
                task_type_status=Substr('reversed_status', TaskType.EMBEDDING.value,
                                        1),
            ).filter(task_type_status__in=state_list, document_id=document_id)

            paragraph_count = paragraph_queryset.count()
            logger.info(f"📊 符合条件的段落数量: {paragraph_count}")

            if paragraph_queryset.exists():
                logger.info("🔄 更新段落向量化状态为PENDING...")
                ListenerManagement.update_status(paragraph_queryset.values('id'),
                                                 TaskType.EMBEDDING,
                                                 State.PENDING)

            logger.info("📊 聚合文档状态...")
            ListenerManagement.get_aggregation_document_status(document_id)()

            try:
                logger.info(f"🚀 提交向量化任务到Celery队列，文档ID: {document_id}, 模型ID: {embedding_model_id}")
                embedding_by_document.delay(document_id, embedding_model_id, state_list)
                logger.info("✅ 向量化任务提交成功")
                logger.info(
                    f"=== 文档向量化刷新调试信息结束 [{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ===")
            except AlreadyQueued as e:
                logger.warning("⚠️ 任务正在执行中，请勿重复提交")
                raise AppApiException(500, _('The task is being executed, please do not send it repeatedly.'))
            except Exception as e:
                logger.error(f"❌ 启动向量化任务失败: {str(e)}")
                # 添加更详细的错误处理
                raise AppApiException(500, _('Failed to start vectorization task: ') + str(e))

        def cancel(self, instance, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                CancelInstanceSerializer(data=instance).is_valid()
            document_id = self.data.get("document_id")
            ListenerManagement.update_status(QuerySet(Paragraph).annotate(
                reversed_status=Reverse('status'),
                task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value,
                                        1),
            ).filter(task_type_status__in=[State.PENDING.value, State.STARTED.value]).filter(
                document_id=document_id).values('id'),
                                             TaskType(instance.get('type')),
                                             State.REVOKE)
            ListenerManagement.update_status(QuerySet(Document).annotate(
                reversed_status=Reverse('status'),
                task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value,
                                        1),
            ).filter(task_type_status__in=[State.PENDING.value, State.STARTED.value]).filter(
                id=document_id).values('id'),
                                             TaskType(instance.get('type')),
                                             State.REVOKE)

            return True

        @transaction.atomic
        def delete(self):
            document_id = self.data.get("document_id")
            QuerySet(model=Document).filter(id=document_id).delete()
            # 删除段落
            QuerySet(model=Paragraph).filter(document_id=document_id).delete()
            # 删除问题
            delete_problems_and_mappings([document_id])
            # 删除向量库
            delete_embedding_by_document(document_id)
            return True

        @staticmethod
        def get_response_body_api():
            return openapi.Schema(
                type=openapi.TYPE_OBJECT,
                required=['id', 'name', 'char_length', 'user_id', 'paragraph_count', 'is_active'
                                                                                     'update_time', 'create_time'],
                properties={
                    'id': openapi.Schema(type=openapi.TYPE_STRING, title="id",
                                         description="id", default="xx"),
                    'name': openapi.Schema(type=openapi.TYPE_STRING, title=_('name'),
                                           description=_('name'), default="xx"),
                    'char_length': openapi.Schema(type=openapi.TYPE_INTEGER, title=_('char length'),
                                                  description=_('char length'), default=10),
                    'user_id': openapi.Schema(type=openapi.TYPE_STRING, title=_('user id'), description=_('user id')),
                    'paragraph_count': openapi.Schema(type=openapi.TYPE_INTEGER, title="_('document count')",
                                                      description="_('document count')", default=1),
                    'is_active': openapi.Schema(type=openapi.TYPE_BOOLEAN, title=_('Is active'),
                                                description=_('Is active'), default=True),
                    'update_time': openapi.Schema(type=openapi.TYPE_STRING, title=_('update time'),
                                                  description=_('update time'),
                                                  default="1970-01-01 00:00:00"),
                    'create_time': openapi.Schema(type=openapi.TYPE_STRING, title=_('create time'),
                                                  description=_('create time'),
                                                  default="1970-01-01 00:00:00"
                                                  )
                }
            )

        @staticmethod
        def get_request_body_api():
            return openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'name': openapi.Schema(type=openapi.TYPE_STRING, title=_('document name'),
                                           description=_('document name')),
                    'is_active': openapi.Schema(type=openapi.TYPE_BOOLEAN, title=_('Is active'),
                                                description=_('Is active')),
                    'hit_handling_method': openapi.Schema(type=openapi.TYPE_STRING, title=_('hit handling method'),
                                                          description=_(
                                                              'ai optimization: optimization, direct return: directly_return')),
                    'directly_return_similarity': openapi.Schema(type=openapi.TYPE_NUMBER,
                                                                 title=_('directly return similarity'),
                                                                 default=0.9),
                    'meta': openapi.Schema(type=openapi.TYPE_OBJECT, title=_('meta'),
                                           description=_(
                                               'Document metadata->web:{source_url:xxx,selector:\'xxx\'},base:{}')),
                }
            )

    class Create(ApiMixin, serializers.Serializer):
        dataset_id = serializers.UUIDField(required=True, error_messages=ErrMessage.char(
            _('document id')))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            if not QuerySet(DataSet).filter(id=self.data.get('dataset_id')).exists():
                raise AppApiException(10000, _('dataset id not exist'))
            return True

        @staticmethod
        def post_embedding(result, document_id, dataset_id):
            DocumentSerializers.Operate(
                data={'dataset_id': dataset_id, 'document_id': document_id}).refresh()
            return result

        @staticmethod
        def parse_qa_file(file):
            get_buffer = FileBufferHandle().get_buffer
            for parse_qa_handle in parse_qa_handle_list:
                if parse_qa_handle.support(file, get_buffer):
                    return parse_qa_handle.handle(file, get_buffer, save_image)
            raise AppApiException(500, _('Unsupported file format'))

        @staticmethod
        def parse_table_file(file):
            get_buffer = FileBufferHandle().get_buffer
            for parse_table_handle in parse_table_handle_list:
                if parse_table_handle.support(file, get_buffer):
                    return parse_table_handle.handle(file, get_buffer, save_image)
            raise AppApiException(500, _('Unsupported file format'))

        @staticmethod
        def parse_audio_file(file):
            """处理音频文件：语音转文字 + 智能分段"""
            from setting.models_provider.tools import get_model_instance_by_model_user_id
            from setting.models_provider.impl.base_stt import BaseSpeechToText
            from common.util.split_model import SplitModel
            import re
            import logging

            # 获取日志记录器
            logger = logging.getLogger('dataset.serializers.document_serializers')

            # 检查文件是否为音频格式
            file_name = file.name.lower()
            audio_extensions = ['.mp3', '.wav', '.ogg', '.aac', '.m4a', '.flac']
            if not any(file_name.endswith(ext) for ext in audio_extensions):
                raise AppApiException(500, _('Unsupported audio file format'))

            import datetime
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            logger.info(f"=== 开始处理音频文件调试信息 [{timestamp}] ===")
            logger.info(f"📁 音频文件名: {file.name}")
            logger.info(f"📊 音频文件大小: {file.size:,} bytes ({file.size / 1024 / 1024:.2f} MB)")

            try:
                # 获取可用的STT模型
                from setting.models import Model
                from django.db.models import Q
                stt_models = Model.objects.filter(
                    Q(model_type='STT') & Q(status='SUCCESS')
                ).order_by('create_time')

                if not stt_models.exists():
                    logger.error("没有找到可用的STT模型")
                    raise AppApiException(500, _('No available STT model found'))

                stt_model = stt_models.first()
                logger.info(f"🎯 使用的STT模型ID: {stt_model.id}")
                logger.info(f"🏷️ STT模型名称: {stt_model.name}")

                # 获取STT模型实例
                model_instance = get_model_instance_by_model_user_id(stt_model.id, None)
                if not isinstance(model_instance, BaseSpeechToText):
                    logger.error("STT模型实例不可用")
                    raise AppApiException(500, _('STT model is not available'))

                logger.info("✅ STT模型实例获取成功，开始语音转文字...")

                # 语音转文字
                start_time = datetime.datetime.now()
                transcribed_text = model_instance.speech_to_text(file)
                end_time = datetime.datetime.now()
                duration = (end_time - start_time).total_seconds()

                logger.info(f"🎤 语音转文字完成，耗时: {duration:.2f}秒")
                logger.info(f"📝 原始结果长度: {len(transcribed_text) if transcribed_text else 0} 字符")
                logger.info(f"📄 语音转文字结果: {transcribed_text}")

                if not transcribed_text or not transcribed_text.strip():
                    logger.error("语音转文字失败或返回空结果")
                    raise AppApiException(500, _('Audio transcription failed or returned empty result'))

                # 智能分段
                text_content = transcribed_text.strip()
                logger.info(f"🔀 开始智能分段处理，文本内容长度: {len(text_content)} 字符")
                logger.info(f"📋 分段前文本内容: {text_content}")

                # 使用默认的分段模式
                default_pattern_list = [
                    re.compile('(?<=^)# .*|(?<=\\n)# .*'),
                    re.compile('(?<=\\n)(?<!#)## (?!#).*|(?<=^)(?<!#)## (?!#).*'),
                    re.compile("(?<=\\n)(?<!#)### (?!#).*|(?<=^)(?<!#)### (?!#).*"),
                    re.compile("(?<=\\n)(?<!#)#### (?!#).*|(?<=^)(?<!#)#### (?!#).*"),
                    re.compile("(?<=\\n)(?<!#)##### (?!#).*|(?<=^)(?<!#)##### (?!#).*"),
                    re.compile("(?<=\\n)(?<!#)###### (?!#).*|(?<=^)(?<!#)###### (?!#).*")
                ]

                split_model = SplitModel(default_pattern_list, with_filter=True, limit=800)
                segments = split_model.parse(text_content)
                logger.info(f"✅ 智能分段完成，分段数量: {len(segments) if segments else 0}")

                if segments:
                    for i, segment in enumerate(segments):
                        logger.info(f"📑 分段 {i + 1}:")
                        logger.info(f"  📌 标题: {segment.get('title', '无标题')}")
                        logger.info(f"  📏 内容长度: {len(segment.get('content', ''))} 字符")
                        logger.info(f"  👀 内容预览: {segment.get('content', '')[:200]}...")
                else:
                    logger.info("⚠️ 没有检测到分段，将创建默认分段")

                if not segments:
                    # 如果没有分段，创建一个默认分段
                    segments = [{'title': '音频转录内容', 'content': text_content}]
                    logger.info("✅ 创建默认分段完成")

                # 返回文档格式
                import re
                file_name_without_ext = re.sub(r'\.[^/.]+$', '', file.name)
                result = [{
                    'name': file_name_without_ext + '_音频转录',
                    'paragraphs': segments
                }]

                logger.info(f"🎉 音频文件处理完成，返回文档数量: {len(result)}")
                logger.info(
                    f"=== 音频文件处理调试信息结束 [{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ===")

                return result

            except Exception as e:
                import traceback
                logger.error(f"音频文件处理失败: {str(e)}")
                logger.error(f"错误堆栈: {traceback.format_exc()}")
                traceback.print_exc()
                raise AppApiException(500, f'Audio processing failed: {str(e)}')

        def save_qa(self, instance: Dict, with_valid=True):
            if with_valid:
                DocumentInstanceQASerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            file_list = instance.get('file_list')
            document_list = flat_map([self.parse_qa_file(file) for file in file_list])
            return DocumentSerializers.Batch(data={'dataset_id': self.data.get('dataset_id')}).batch_save(document_list)

        def save_table(self, instance: Dict, with_valid=True):
            if with_valid:
                DocumentInstanceTableSerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            file_list = instance.get('file_list')
            document_list = flat_map([self.parse_table_file(file) for file in file_list])
            return DocumentSerializers.Batch(data={'dataset_id': self.data.get('dataset_id')}).batch_save(document_list)

        def save_audio(self, instance: Dict, with_valid=True):
            if with_valid:
                DocumentInstanceQASerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            file_list = instance.get('file_list')
            document_list = flat_map([self.parse_audio_file(file) for file in file_list])
            return DocumentSerializers.Batch(data={'dataset_id': self.data.get('dataset_id')}).batch_save(document_list)

        def save_data_source(self, dataset_id, data: Dict):
            source_id = data.get('source_id')
            table_name = data.get('table_name')
            columns = data.get('columns')
            if not columns:
                raise AppApiException(500, _('Please select database fields'))
            data_source = DataSourceConfig.objects.filter(id=source_id).first()
            params = {
                "db_type": data_source.db_type,
                "user": data_source.username,
                "password": data_source.password,
                "host": data_source.host,
                "port": data_source.port,
                "dbname": data_source.database_name,
            }

            # 处理Oracle特有的参数
            if data_source.db_type == 'oracle':
                extra_params = data_source.extra_params or {}
                oracle_connect_type = extra_params.get('oracle_connect_type', 'sid')

                if oracle_connect_type == 'sid':
                    params['sid'] = data_source.database_name
                else:
                    params['service_name'] = data_source.database_name

            # 添加schema参数（如果存在）
            if data_source.extra_params and data_source.extra_params.get('schema'):
                params['schema'] = data_source.extra_params.get('schema')

            columns_list = DBConnector.query_columns(params, table_name, columns)
            if not columns_list:
                raise AppApiException(500, _('The data table has no data'))
            content_list = [
                {'name': table_name, 'paragraphs': [{"title": "", "content": str(item)} for item in columns_list]}]
            return DocumentSerializers.Batch(data={'dataset_id': dataset_id}).batch_save(content_list)

        @post(post_function=post_embedding)
        @transaction.atomic
        def save(self, instance: Dict, with_valid=False, **kwargs):
            if with_valid:
                DocumentInstanceSerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            dataset_id = self.data.get('dataset_id')
            document_paragraph_model = self.get_document_paragraph_model(dataset_id, instance)

            document_model = document_paragraph_model.get('document')
            paragraph_model_list = document_paragraph_model.get('paragraph_model_list')
            problem_paragraph_object_list = document_paragraph_model.get('problem_paragraph_object_list')
            problem_model_list, problem_paragraph_mapping_list = (ProblemParagraphManage(problem_paragraph_object_list,
                                                                                         dataset_id)
                                                                  .to_problem_model_list())
            # 插入文档
            document_model.save()
            # 批量插入段落
            QuerySet(Paragraph).bulk_create(paragraph_model_list) if len(paragraph_model_list) > 0 else None
            # 批量插入问题
            QuerySet(Problem).bulk_create(problem_model_list) if len(problem_model_list) > 0 else None
            # 批量插入关联问题
            QuerySet(ProblemParagraphMapping).bulk_create(problem_paragraph_mapping_list) if len(
                problem_paragraph_mapping_list) > 0 else None
            document_id = str(document_model.id)
            return DocumentSerializers.Operate(
                data={'dataset_id': dataset_id, 'document_id': document_id}).one(
                with_valid=True), document_id, dataset_id

        def save_web(self, instance: Dict, with_valid=True):
            if with_valid:
                DocumentWebInstanceSerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            dataset_id = self.data.get('dataset_id')
            source_url_list = instance.get('source_url_list')
            selector = instance.get('selector')
            sync_web_document.delay(dataset_id, source_url_list, selector)

        @staticmethod
        def get_paragraph_model(document_model, paragraph_list: List):
            dataset_id = document_model.dataset_id
            paragraph_model_dict_list = [ParagraphSerializers.Create(
                data={'dataset_id': dataset_id, 'document_id': str(document_model.id)}).get_paragraph_problem_model(
                dataset_id, document_model.id, paragraph) for paragraph in paragraph_list]

            paragraph_model_list = []
            problem_paragraph_object_list = []
            for paragraphs in paragraph_model_dict_list:
                paragraph = paragraphs.get('paragraph')
                for problem_model in paragraphs.get('problem_paragraph_object_list'):
                    problem_paragraph_object_list.append(problem_model)
                paragraph_model_list.append(paragraph)

            return {'document': document_model, 'paragraph_model_list': paragraph_model_list,
                    'problem_paragraph_object_list': problem_paragraph_object_list}

        @staticmethod
        def get_document_paragraph_model(dataset_id, instance: Dict):
            document_model = Document(
                **{'dataset_id': dataset_id,
                   'id': uuid.uuid1(),
                   'name': instance.get('name'),
                   'char_length': reduce(lambda x, y: x + y,
                                         [len(p.get('content')) for p in instance.get('paragraphs', [])],
                                         0),
                   'meta': instance.get('meta') if instance.get('meta') is not None else {},
                   'type': instance.get('type') if instance.get('type') is not None else Type.base})

            return DocumentSerializers.Create.get_paragraph_model(document_model,
                                                                  instance.get('paragraphs') if
                                                                  'paragraphs' in instance else [])

        @staticmethod
        def get_request_body_api():
            return DocumentInstanceSerializer.get_request_body_api()

        @staticmethod
        def get_request_params_api():
            return [openapi.Parameter(name='dataset_id',
                                      in_=openapi.IN_PATH,
                                      type=openapi.TYPE_STRING,
                                      required=True,
                                      description=_('document id'))
                    ]

    class Split(ApiMixin, serializers.Serializer):
        file = serializers.ListField(required=True, error_messages=ErrMessage.list(
            _('file list')))

        limit = serializers.IntegerField(required=False, error_messages=ErrMessage.integer(
            _('limit')))

        patterns = serializers.ListField(required=False,
                                         child=serializers.CharField(required=True, error_messages=ErrMessage.char(
                                             _('patterns'))),
                                         error_messages=ErrMessage.list(
                                             _('patterns')))

        with_filter = serializers.BooleanField(required=False, error_messages=ErrMessage.boolean(
            _('Auto Clean')))

        def is_valid(self, *, raise_exception=True):
            super().is_valid(raise_exception=True)
            files = self.data.get('file')
            for f in files:
                if f.size > 1024 * 1024 * 100:
                    raise AppApiException(500, _('The maximum size of the uploaded file cannot exceed 100MB'))

        @staticmethod
        def get_request_params_api():
            return [
                openapi.Parameter(name='file',
                                  in_=openapi.IN_FORM,
                                  type=openapi.TYPE_ARRAY,
                                  items=openapi.Items(type=openapi.TYPE_FILE),
                                  required=True,
                                  description=_('file list')),
                openapi.Parameter(name='limit',
                                  in_=openapi.IN_FORM,
                                  required=False,
                                  type=openapi.TYPE_INTEGER, title=_('limit'), description=_('limit')),
                openapi.Parameter(name='patterns',
                                  in_=openapi.IN_FORM,
                                  required=False,
                                  type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                                  title=_('Segmented regular list'), description=_('Segmented regular list')),
                openapi.Parameter(name='with_filter',
                                  in_=openapi.IN_FORM,
                                  required=False,
                                  type=openapi.TYPE_BOOLEAN, title=_('Whether to clear special characters'),
                                  description=_('Whether to clear special characters')),
            ]

        def parse(self):
            file_list = self.data.get("file")
            return reduce(lambda x, y: [*x, *y],
                          [file_to_paragraph(f, self.data.get("patterns", None), self.data.get("with_filter", None),
                                             self.data.get("limit", 4096)) for f in file_list], [])

    class SplitPattern(ApiMixin, serializers.Serializer):
        @staticmethod
        def list():
            return [{'key': "#", 'value': '(?<=^)# .*|(?<=\\n)# .*'},
                    {'key': '##', 'value': '(?<=\\n)(?<!#)## (?!#).*|(?<=^)(?<!#)## (?!#).*'},
                    {'key': '###', 'value': "(?<=\\n)(?<!#)### (?!#).*|(?<=^)(?<!#)### (?!#).*"},
                    {'key': '####', 'value': "(?<=\\n)(?<!#)#### (?!#).*|(?<=^)(?<!#)#### (?!#).*"},
                    {'key': '#####', 'value': "(?<=\\n)(?<!#)##### (?!#).*|(?<=^)(?<!#)##### (?!#).*"},
                    {'key': '######', 'value': "(?<=\\n)(?<!#)###### (?!#).*|(?<=^)(?<!#)###### (?!#).*"},
                    {'key': '-', 'value': '(?<! )- .*'},
                    {'key': _('space'), 'value': '(?<! ) (?! )'},
                    {'key': _('semicolon'), 'value': '(?<!；)；(?!；)'}, {'key': _('comma'), 'value': '(?<!，)，(?!，)'},
                    {'key': _('period'), 'value': '(?<!。)。(?!。)'}, {'key': _('enter'), 'value': '(?<!\\n)\\n(?!\\n)'},
                    {'key': _('blank line'), 'value': '(?<!\\n)\\n\\n(?!\\n)'}]

    class Batch(ApiMixin, serializers.Serializer):
        dataset_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_('dataset id')))

        @staticmethod
        def get_request_body_api():
            return openapi.Schema(type=openapi.TYPE_ARRAY, items=DocumentSerializers.Create.get_request_body_api())

        @staticmethod
        def post_embedding(document_list, dataset_id):
            for document_dict in document_list:
                DocumentSerializers.Operate(
                    data={'dataset_id': dataset_id, 'document_id': document_dict.get('id')}).refresh()
            return document_list

        @post(post_function=post_embedding)
        @transaction.atomic
        def batch_save(self, instance_list: List[Dict], with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            DocumentInstanceSerializer(many=True, data=instance_list).is_valid(raise_exception=True)
            dataset_id = self.data.get("dataset_id")
            document_model_list = []
            paragraph_model_list = []
            problem_paragraph_object_list = []
            # 插入文档
            for document in instance_list:
                document_paragraph_dict_model = DocumentSerializers.Create.get_document_paragraph_model(dataset_id,
                                                                                                        document)
                document_model_list.append(document_paragraph_dict_model.get('document'))
                for paragraph in document_paragraph_dict_model.get('paragraph_model_list'):
                    paragraph_model_list.append(paragraph)
                for problem_paragraph_object in document_paragraph_dict_model.get('problem_paragraph_object_list'):
                    problem_paragraph_object_list.append(problem_paragraph_object)

            problem_model_list, problem_paragraph_mapping_list = (ProblemParagraphManage(problem_paragraph_object_list,
                                                                                         dataset_id)
                                                                  .to_problem_model_list())
            # 插入文档
            QuerySet(Document).bulk_create(document_model_list) if len(document_model_list) > 0 else None
            # 批量插入段落
            bulk_create_in_batches(Paragraph, paragraph_model_list, batch_size=1000)
            # 批量插入问题
            bulk_create_in_batches(Problem, problem_model_list, batch_size=1000)
            # 批量插入关联问题
            bulk_create_in_batches(ProblemParagraphMapping, problem_paragraph_mapping_list, batch_size=1000)
            # 查询文档
            query_set = QuerySet(model=Document)
            if len(document_model_list) == 0:
                return [], dataset_id
            query_set = query_set.filter(**{'id__in': [d.id for d in document_model_list]})
            return native_search({
                'document_custom_sql': query_set,
                'order_by_query': QuerySet(Document).order_by('-create_time', 'id')
            }, select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "dataset", 'sql', 'list_document.sql')),
                with_search_one=False), dataset_id

        @staticmethod
        def _batch_sync(document_id_list: List[str]):
            for document_id in document_id_list:
                DocumentSerializers.Sync(data={'document_id': document_id}).sync()

        def batch_sync(self, instance: Dict, with_valid=True):
            if with_valid:
                BatchSerializer(data=instance).is_valid(model=Document, raise_exception=True)
                self.is_valid(raise_exception=True)
            # 异步同步
            work_thread_pool.submit(self._batch_sync,
                                    instance.get('id_list'))
            return True

        @transaction.atomic
        def batch_delete(self, instance: Dict, with_valid=True):
            if with_valid:
                BatchSerializer(data=instance).is_valid(model=Document, raise_exception=True)
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            QuerySet(Document).filter(id__in=document_id_list).delete()
            QuerySet(Paragraph).filter(document_id__in=document_id_list).delete()
            delete_problems_and_mappings(document_id_list)
            # 删除向量库
            delete_embedding_by_document_list(document_id_list)
            return True

        def batch_cancel(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                BatchCancelInstanceSerializer(data=instance).is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            ListenerManagement.update_status(QuerySet(Paragraph).annotate(
                reversed_status=Reverse('status'),
                task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value,
                                        1),
            ).filter(task_type_status__in=[State.PENDING.value, State.STARTED.value]).filter(
                document_id__in=document_id_list).values('id'),
                                             TaskType(instance.get('type')),
                                             State.REVOKE)
            ListenerManagement.update_status(QuerySet(Document).annotate(
                reversed_status=Reverse('status'),
                task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value,
                                        1),
            ).filter(task_type_status__in=[State.PENDING.value, State.STARTED.value]).filter(
                id__in=document_id_list).values('id'),
                                             TaskType(instance.get('type')),
                                             State.REVOKE)

        def batch_edit_hit_handling(self, instance: Dict, with_valid=True):
            if with_valid:
                BatchSerializer(data=instance).is_valid(model=Document, raise_exception=True)
                hit_handling_method = instance.get('hit_handling_method')
                if hit_handling_method is None:
                    raise AppApiException(500, _('Hit handling method is required'))
                if hit_handling_method != 'optimization' and hit_handling_method != 'directly_return':
                    raise AppApiException(500, _('The hit processing method must be directly_return|optimization'))
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            hit_handling_method = instance.get('hit_handling_method')
            directly_return_similarity = instance.get('directly_return_similarity')
            update_dict = {'hit_handling_method': hit_handling_method}
            if directly_return_similarity is not None:
                update_dict['directly_return_similarity'] = directly_return_similarity
            QuerySet(Document).filter(id__in=document_id_list).update(**update_dict)

        def batch_refresh(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            state_list = instance.get("state_list")
            dataset_id = self.data.get('dataset_id')
            for document_id in document_id_list:
                try:
                    DocumentSerializers.Operate(
                        data={'dataset_id': dataset_id, 'document_id': document_id}).refresh(state_list)
                except AlreadyQueued as e:
                    pass

    class GenerateRelated(ApiMixin, serializers.Serializer):
        document_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_('document id')))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            document_id = self.data.get('document_id')
            if not QuerySet(Document).filter(id=document_id).exists():
                raise AppApiException(500, _('document id not exist'))

        def generate_related(self, model_id, prompt, state_list=None, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_id = self.data.get('document_id')
            ListenerManagement.update_status(QuerySet(Document).filter(id=document_id),
                                             TaskType.GENERATE_PROBLEM,
                                             State.PENDING)
            ListenerManagement.update_status(QuerySet(Paragraph).filter(document_id=document_id),
                                             TaskType.GENERATE_PROBLEM,
                                             State.PENDING)
            ListenerManagement.get_aggregation_document_status(document_id)()
            try:
                generate_related_by_document_id.delay(document_id, model_id, prompt, state_list)
            except AlreadyQueued as e:
                raise AppApiException(500, _('The task is being executed, please do not send it again.'))

    class BatchGenerateRelated(ApiMixin, serializers.Serializer):
        dataset_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_('dataset id')))

        def batch_generate_related(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("document_id_list")
            model_id = instance.get("model_id")
            prompt = instance.get("prompt")
            state_list = instance.get('state_list')
            ListenerManagement.update_status(QuerySet(Document).filter(id__in=document_id_list),
                                             TaskType.GENERATE_PROBLEM,
                                             State.PENDING)
            ListenerManagement.update_status(QuerySet(Paragraph).annotate(
                reversed_status=Reverse('status'),
                task_type_status=Substr('reversed_status', TaskType.GENERATE_PROBLEM.value,
                                        1),
            ).filter(task_type_status__in=state_list, document_id__in=document_id_list)
                                             .values('id'),
                                             TaskType.GENERATE_PROBLEM,
                                             State.PENDING)
            ListenerManagement.get_aggregation_document_status_by_query_set(
                QuerySet(Document).filter(id__in=document_id_list))()
            try:
                for document_id in document_id_list:
                    generate_related_by_document_id.delay(document_id, model_id, prompt, state_list)
            except AlreadyQueued as e:
                pass


class FileBufferHandle:
    buffer = None

    def get_buffer(self, file):
        if self.buffer is None:
            self.buffer = file.read()
        return self.buffer


default_split_handle = TextSplitHandle()
split_handles = [HTMLSplitHandle(), DocSplitHandle(), PdfSplitHandle(), XlsxSplitHandle(), XlsSplitHandle(),
                 CsvSplitHandle(),
                 ZipSplitHandle(),
                 default_split_handle]


def save_image(image_list):
    if image_list is not None and len(image_list) > 0:
        exist_image_list = [str(i.get('id')) for i in
                            QuerySet(Image).filter(id__in=[i.id for i in image_list]).values('id')]
        save_image_list = [image for image in image_list if not exist_image_list.__contains__(str(image.id))]
        save_image_list = list({img.id: img for img in save_image_list}.values())
        if len(save_image_list) > 0:
            QuerySet(Image).bulk_create(save_image_list)


def file_to_paragraph(file, pattern_list: List, with_filter: bool, limit: int):
    get_buffer = FileBufferHandle().get_buffer
    for split_handle in split_handles:
        if split_handle.support(file, get_buffer):
            result = split_handle.handle(file, pattern_list, with_filter, limit, get_buffer, save_image)
            if isinstance(result, list):
                return result
            return [result]
    result = default_split_handle.handle(file, pattern_list, with_filter, limit, get_buffer, save_image)
    if isinstance(result, list):
        return result
    return [result]


def delete_problems_and_mappings(document_ids):
    problem_paragraph_mappings = ProblemParagraphMapping.objects.filter(document_id__in=document_ids)
    problem_ids = set(problem_paragraph_mappings.values_list('problem_id', flat=True))

    if problem_ids:
        problem_paragraph_mappings.delete()
        remaining_problem_counts = ProblemParagraphMapping.objects.filter(problem_id__in=problem_ids).values(
            'problem_id').annotate(count=Count('problem_id'))
        remaining_problem_ids = {pc['problem_id'] for pc in remaining_problem_counts}
        problem_ids_to_delete = problem_ids - remaining_problem_ids
        Problem.objects.filter(id__in=problem_ids_to_delete).delete()
    else:
        problem_paragraph_mappings.delete()
